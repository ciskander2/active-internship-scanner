import argparse
import datetime as dt
import html
import json
import os
import re
import shutil
import time
import urllib.parse
import urllib.request
from urllib.error import HTTPError, URLError

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROVIDERS = [
    ("greenhouse", "https://boards-api.greenhouse.io/v1/boards/{slug}/jobs?content=true"),
    ("lever", "https://api.lever.co/v0/postings/{slug}?mode=json"),
    ("ashby", "https://api.ashbyhq.com/posting-api/job-board/{slug}"),
]

OFFICIAL_SOURCE_DOMAINS = {
    "Amazon": ["amazon.jobs"],
    "Adobe": ["adobe.wd5.myworkdayjobs.com"],
    "DE Shaw": ["deshaw.com"],
    "Intuit": ["jobs.intuit.com"],
    "Netflix": ["explore.jobs.netflix.net", "jobs.netflix.com"],
    "NVIDIA": ["nvidia.wd5.myworkdayjobs.com"],
    "Optiver": ["optiver.com"],
    "Rivian": ["careers.rivian.com"],
    "Salesforce": ["salesforce.wd12.myworkdayjobs.com"],
    "Susquehanna Associates": ["careers.sig.com", "sig.com"],
    "Tesla": ["tesla.com"],
}

OFFICIAL_CAREER_PAGES = {
    "DE Shaw": ["https://www.deshaw.com/careers?keywords=intern"],
    "Intuit": ["https://jobs.intuit.com/search-jobs/intern/27595/1"],
    "Netflix": ["https://explore.jobs.netflix.net/careers?query=intern"],
    "Optiver": [
        "https://www.optiver.com/working-at-optiver/career-opportunities/?level=internship&numberposts=100"
    ],
    "Tesla": ["https://www.tesla.com/careers/search/?query=internship"],
}

WORKDAY_SOURCES = {
    "Adobe": [
        ("https://adobe.wd5.myworkdayjobs.com", "adobe", "external_experienced"),
        ("https://adobe.wd5.myworkdayjobs.com", "adobe", "external_university"),
    ],
    "NVIDIA": [
        ("https://nvidia.wd5.myworkdayjobs.com", "nvidia", "NVIDIAExternalCareerSite"),
    ],
    "Salesforce": [
        ("https://salesforce.wd12.myworkdayjobs.com", "salesforce", "External_Career_Site"),
    ],
}

SLUGS = {
    "Microsoft": ["microsoft"],
    "Bloomberg": ["bloomberg"],
    "Meta": ["meta"],
    "Figma": ["figma"],
    "Dropbox": ["dropbox"],
    "Snowflake": ["snowflake"],
    "Adobe": ["adobe"],
    "Intuit": ["intuit"],
    "Airbnb": ["airbnb"],
    "Spotify": ["spotify"],
    "Uber": ["uber"],
    "Salesforce": ["salesforce"],
    "Lyft": ["lyft"],
    "Amazon": ["amazon"],
    "MongoDb": ["mongodb"],
    "Coinbase": ["coinbase"],
    "Atlassian": ["atlassian"],
    "Shopify": ["shopify"],
    "Pinterest": ["pinterest"],
    "Point 72": ["point72"],
    "SeatGeek": ["seatgeek"],
    "TikTok": ["bytedance", "tiktok"],
    "Databricks": ["databricks"],
    "Duolingo": ["duolingo"],
    "Cloudflare": ["cloudflare"],
    "Tesla": ["tesla"],
    "SoFi": ["sofi"],
    "Reddit": ["reddit"],
    "Two Sigma": ["twosigma"],
    "Rivian": ["rivian"],
    "Millenium": ["millennium"],
    "Citadel": ["citadel", "citadel-securities"],
    "SpaceX": ["spacex"],
    "Gusto": ["gusto"],
    "Snap Inc": ["snapchat", "snap"],
    "Zendesk": ["zendesk"],
    "Netflix": ["netflix"],
    "Precisely": ["precisely"],
    "Confluent": ["confluent"],
    "Robinhood": ["robinhood"],
    "NVIDIA": ["nvidia"],
    "Palantir": ["palantir"],
    "Anduril Industries": ["andurilindustries", "anduril"],
    "Stripe": ["stripe"],
    "Ramp": ["ramp"],
    "ServiceNow": ["servicenow"],
    "Discord": ["discord"],
    "Notion": ["notion"],
    "xAI": ["xai"],
    "Mercury": ["mercury"],
    "DoorDash": ["doordash"],
    "Google": ["google"],
    "Volirdge": ["voloridgeinvestmentmanagement", "voloridge"],
    "Susquehanna Associates": ["sig", "susquehanna"],
    "Bridgewater Associates": ["bridgewater"],
    "DE Shaw": ["deshaw", "the-d-e-shaw-group", "d-e-shaw"],
    "Hudson River Trading": ["hudsonrivertrading", "hrt"],
    "Akuna Capital": ["akunacapital", "akuna"],
    "IMC Trading": ["imc", "imc-trading"],
    "Optiver": ["optiver"],
    "Jump Trading": ["jumptrading", "jump-trading"],
    "Tower Research Capital": ["towerresearchcapital", "tower-research", "towerresearch"],
    "Five Rings": ["fiverings", "five-rings"],
    "Belvedere Trading": ["belvederetrading", "belvedere-trading"],
    "Old Mission": ["oldmission", "old-mission"],
    "Flow Trader": ["flowtraders", "flow-traders"],
    "Geneva Trading": ["genevatrading", "geneva-trading"],
    "XR Trading": ["xrtrading", "xr-trading"],
    "AQR Capital Management": ["aqr", "aqr-capital-management"],
    "Man Group": ["mangroup", "man-group"],
    "Balyasny Asset Management": ["balyasny", "bamfunds"],
    "Schonfeld Strategic Advisors": ["schonfeld", "schonfeldstrategicadvisors"],
    "Datadog": ["datadog"],
    "Jane Street": ["janestreet", "jane-street"],
}

def normalize_company(company):
    return re.sub(r"[^a-z0-9]", "", company.lower().replace("&", "and"))


def variants(company):
    values = list(SLUGS.get(company, []))
    base = normalize_company(company)
    values.extend(
        [
            base,
            base.replace("inc", ""),
            base.replace("industries", ""),
            base.replace("capitalmanagement", ""),
            base.replace("trading", ""),
        ]
    )
    out = []
    for value in values:
        value = value.strip("-")
        if value and value not in out:
            out.append(value)
    return out[:8]


def fetch_json(url, timeout=15):
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"},
    )
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8", "ignore"))


def fetch_text(url, timeout=15):
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "Mozilla/5.0", "Accept": "text/html,application/xhtml+xml"},
    )
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return response.read().decode("utf-8", "ignore")


def clean_html(value):
    value = re.sub(r"<.*?>", " ", value or "", flags=re.S)
    return re.sub(r"\s+", " ", html.unescape(value)).strip()


def official_domains(company):
    domains = list(OFFICIAL_SOURCE_DOMAINS.get(company, []))
    for slug in variants(company):
        if "." in slug:
            domains.append(slug)
    return list(dict.fromkeys(domains))


def absolute_url(url, base_url):
    return urllib.parse.urljoin(base_url, html.unescape(url))


def collect_json_jobs(value, base_url):
    jobs = []
    if isinstance(value, dict):
        title = value.get("title") or value.get("name")
        url = value.get("url") or value.get("href")
        if title and url and is_active_internship({"title": title}):
            jobs.append(
                {
                    "title": clean_html(str(title)),
                    "url": absolute_url(str(url), base_url),
                    "location": clean_html(text_of(value.get("location") or "")),
                    "body": clean_html(text_of(value)),
                    "source": "official_source",
                }
            )
        for nested in value.values():
            jobs.extend(collect_json_jobs(nested, base_url))
    elif isinstance(value, list):
        for nested in value:
            jobs.extend(collect_json_jobs(nested, base_url))
    return jobs


def parse_official_html_jobs(text, base_url):
    jobs = []
    for script in re.findall(
        r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
        text,
        re.S | re.I,
    ):
        try:
            jobs.extend(collect_json_jobs(json.loads(html.unescape(script)), base_url))
        except json.JSONDecodeError:
            continue

    for href, label in re.findall(
        r'<a[^>]+href=["\']([^"\']+)["\'][^>]*>'
        r'(?:(?!</a>).)*?<span[^>]+class=["\'][^"\']*job-display-name[^"\']*["\'][^>]*>(.*?)</span>',
        text,
        re.S | re.I,
    ):
        title = clean_html(label)
        url = absolute_url(href, base_url)
        if not url.startswith(("http://", "https://")):
            continue
        if not is_active_internship({"title": title}):
            continue
        jobs.append(
            {
                "title": title,
                "url": url,
                "location": "",
                "body": title,
                "source": "official_source",
            }
        )

    for href, label in re.findall(r'<a[^>]+href=["\']([^"\']+)["\'][^>]*>(.*?)</a>', text, re.S | re.I):
        title = clean_html(label)
        url = absolute_url(href, base_url)
        if not url.startswith(("http://", "https://")):
            continue
        if not is_active_internship({"title": title}):
            continue
        jobs.append(
            {
                "title": title,
                "url": url,
                "location": "",
                "body": title,
                "source": "official_source",
            }
        )
    return dedupe_jobs(jobs)


def dedupe_jobs(jobs):
    deduped = []
    for job in jobs:
        key = job.get("url") or job.get("title")
        if not key:
            continue
        if not any((existing.get("url") or existing.get("title")) == key for existing in deduped):
            deduped.append(job)
    return deduped


def fetch_amazon_jobs():
    jobs = []
    for offset in range(0, 300, 100):
        url = (
            "https://www.amazon.jobs/en/search.json?"
            f"base_query=internship&offset={offset}&result_limit=100"
        )
        try:
            data = fetch_json(url)
        except (HTTPError, URLError, TimeoutError, json.JSONDecodeError, OSError):
            break
        for job in data.get("jobs", []):
            title = job.get("title", "")
            if not is_active_internship({"title": title}):
                continue
            jobs.append(
                {
                    "title": title,
                    "url": absolute_url(job.get("job_path", ""), "https://www.amazon.jobs"),
                    "location": job.get("location", ""),
                    "body": text_of(job),
                    "source": "official_source",
                }
            )
        if len(data.get("jobs", [])) < 100:
            break
    return dedupe_jobs(jobs)


def fetch_workday_jobs(company):
    jobs = []
    for base_url, tenant, site in WORKDAY_SOURCES.get(company, []):
        endpoint = f"{base_url}/wday/cxs/{tenant}/{site}/jobs"
        for offset in range(0, 100, 20):
            payload = json.dumps(
                {"appliedFacets": {}, "limit": 20, "offset": offset, "searchText": "intern"}
            ).encode("utf-8")
            try:
                req = urllib.request.Request(
                    endpoint,
                    data=payload,
                    headers={
                        "User-Agent": "Mozilla/5.0",
                        "Accept": "application/json",
                        "Content-Type": "application/json",
                    },
                    method="POST",
                )
                with urllib.request.urlopen(req, timeout=20) as response:
                    data = json.loads(response.read().decode("utf-8", "ignore"))
            except (HTTPError, URLError, TimeoutError, json.JSONDecodeError, OSError):
                break

            postings = data.get("jobPostings", [])
            for job in postings:
                title = job.get("title", "")
                if not is_active_internship({"title": title}):
                    continue
                jobs.append(
                    {
                        "title": title,
                        "url": absolute_url(f"/{site}{job.get('externalPath', '')}", base_url),
                        "location": job.get("locationsText", ""),
                        "body": text_of(job),
                        "source": "official_source",
                    }
                )
            if len(postings) < 20:
                break
    return dedupe_jobs(jobs)


def fetch_jibe_jobs(company, base_url):
    jobs = []
    urls = [
        f"{base_url}/api/jobs?keywords=intern&sortBy=relevance&page=1&limit=100",
        f"{base_url}/api/jobs?keywords=internship&sortBy=relevance&page=1&limit=100",
    ]
    if "sig.com" in base_url:
        urls.append(f"{base_url}/api/jobs?categories=Interns%20%2B%20Co-ops&page=1&limit=100")
    for url in urls:
        try:
            data = fetch_json(url)
        except (HTTPError, URLError, TimeoutError, json.JSONDecodeError, OSError):
            continue
        for item in data.get("jobs", []):
            job = item.get("data", item)
            title = job.get("title", "")
            if not is_active_internship({"title": title}):
                continue
            slug = job.get("slug") or job.get("req_id") or ""
            language = job.get("language") or "en-us"
            jobs.append(
                {
                    "title": title,
                    "url": f"{base_url}/jobs/{slug}?lang={language}",
                    "location": job.get("city", ""),
                    "body": text_of(job),
                    "source": "official_source",
                }
            )
    return dedupe_jobs(jobs)


def parse_talentbrew_jobs(text, base_url):
    jobs = []
    pattern = (
        r'<a[^>]+href=["\']([^"\']+)["\'][^>]*'
        r'data-job-id=["\'][^"\']+["\'][^>]*'
        r'data-title=["\']([^"\']+)["\'][^>]*>'
    )
    for href, title in re.findall(pattern, text, re.S | re.I):
        title = clean_html(title)
        if not is_active_internship({"title": title}):
            continue
        jobs.append(
            {
                "title": title,
                "url": absolute_url(href, base_url),
                "location": "",
                "body": title,
                "source": "official_source",
            }
        )
    return dedupe_jobs(jobs)


def parse_eightfold_jobs(text):
    jobs = []
    text = html.unescape(text)
    match = re.search(r'"positions"\s*:\s*(\[.*?\])\s*,\s*"debug"', text, re.S)
    if not match:
        return jobs
    try:
        positions = json.loads(match.group(1))
    except json.JSONDecodeError:
        return jobs

    for job in positions:
        title = job.get("name") or job.get("posting_name") or ""
        if not is_active_internship({"title": title}):
            continue
        jobs.append(
            {
                "title": title,
                "url": job.get("canonicalPositionUrl", ""),
                "location": job.get("location", ""),
                "body": text_of(job),
                "source": "official_source",
            }
        )
    return dedupe_jobs(jobs)


def fetch_official_source_jobs(company):
    workday_jobs = fetch_workday_jobs(company)
    if workday_jobs:
        return workday_jobs
    if company == "Amazon":
        return fetch_amazon_jobs()
    if company == "Susquehanna Associates":
        return fetch_jibe_jobs(company, "https://careers.sig.com")
    if company == "Rivian":
        return fetch_jibe_jobs(company, "https://careers.rivian.com")

    jobs = []
    for url in OFFICIAL_CAREER_PAGES.get(company, []):
        try:
            text = fetch_text(url)
            if company == "Intuit":
                jobs.extend(parse_talentbrew_jobs(text, url))
            elif company == "Netflix":
                jobs.extend(parse_official_html_jobs(text, url))
                jobs.extend(parse_eightfold_jobs(text))
            else:
                jobs.extend(parse_official_html_jobs(text, url))
        except (HTTPError, URLError, TimeoutError, OSError):
            continue
    return dedupe_jobs(jobs)


def text_of(value):
    if isinstance(value, dict):
        return " ".join(text_of(v) for v in value.values())
    if isinstance(value, list):
        return " ".join(text_of(v) for v in value)
    return "" if value is None else str(value)


def extract_jobs(provider, data):
    jobs = []
    if provider == "greenhouse":
        for job in data.get("jobs", []):
            locations = ", ".join(
                location.get("name", "") for location in job.get("locations", [])
            )
            if not locations:
                locations = (job.get("location") or {}).get("name", "")
            jobs.append(
                {
                    "title": job.get("title", ""),
                    "url": job.get("absolute_url", ""),
                    "location": locations,
                    "body": text_of(job),
                }
            )
    elif provider == "lever":
        for job in data:
            categories = job.get("categories", {}) or {}
            jobs.append(
                {
                    "title": job.get("text", ""),
                    "url": job.get("hostedUrl") or job.get("applyUrl") or "",
                    "location": categories.get("location", ""),
                    "body": text_of(job),
                }
            )
    elif provider == "ashby":
        for job in data.get("jobs", []):
            jobs.append(
                {
                    "title": job.get("title", ""),
                    "url": job.get("jobUrl") or job.get("applyUrl") or "",
                    "location": job.get("locationName", ""),
                    "body": text_of(job),
                }
            )
    return jobs


def is_active_internship(job):
    title = (job.get("title") or "").lower()
    internship = r"\binterns?\b|\binternships?\b|\bco-?op\b|\bsummer analyst\b|\bsummer associate\b"
    return bool(re.search(internship, title))


def scan_company(company, delay_seconds=0.1):
    checked = []
    board = None
    all_jobs = []
    official_source_domains = official_domains(company)
    for slug in variants(company):
        for provider, template in PROVIDERS:
            url = template.format(slug=slug)
            checked.append(url)
            try:
                data = fetch_json(url)
                jobs = extract_jobs(provider, data)
            except (HTTPError, URLError, TimeoutError, json.JSONDecodeError, OSError):
                jobs = []
            if jobs:
                board = {"provider": provider, "slug": slug, "url": url}
                all_jobs = jobs
                break
        if all_jobs:
            break
    time.sleep(delay_seconds)
    matches = []
    for job in all_jobs:
        if is_active_internship(job):
            if not any(existing.get("url") == job.get("url") for existing in matches):
                matches.append(job)
    official_source_matches = fetch_official_source_jobs(company)
    for job in official_source_matches:
        if not any(existing.get("url") == job.get("url") for existing in matches):
            matches.append(job)
    return {
        "company": company,
        "board": board,
        "job_count": len(all_jobs),
        "matches": matches,
        "checked": checked,
        "official_source_domains": official_source_domains,
        "official_source_match_count": len(official_source_matches),
    }


def read_companies(workbook_path, sheet_name):
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook[sheet_name]
    companies = []
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row, 1).value
        if value and str(value).strip():
            companies.append((row, str(value).strip()))
    return companies


def style_header(cell):
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=Side(style="thin", color="D9E2F3"))


def write_workbook(input_path, output_path, sheet_name, results, scan_date):
    shutil.copy2(input_path, output_path)
    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook[sheet_name]
    by_company = {result["company"]: result for result in results}

    start_col = 9
    headers = [
        "Active Internship Found?",
        "Internship Role(s) Found",
        "Location(s)",
        "Official Source URL(s)",
        "Scan Notes",
        "Last Checked",
    ]
    for index, header in enumerate(headers, start_col):
        cell = sheet.cell(1, index, header)
        style_header(cell)

    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row, 1).value
        if not value or not str(value).strip():
            continue
        company = str(value).strip()
        result = by_company.get(company, {})
        matches = result.get("matches", [])
        if matches:
            status = "Yes"
            fill = "C6EFCE"
            roles = "; ".join(job.get("title", "") for job in matches[:10])
            locations = "; ".join(
                dict.fromkeys(job.get("location", "") for job in matches[:10] if job.get("location"))
            )
            urls = "; ".join(job.get("url", "") for job in matches[:10] if job.get("url"))
            sources = []
            if any(job.get("source") == "official_source" for job in matches):
                sources.append("official career source")
            if any(job.get("source") != "official_source" for job in matches):
                sources.append("live ATS feed")
            note = (
                f"Confirmed active official internship posting(s) found. "
                f"{len(matches)} match(es) via {', '.join(sources)}."
            )
        else:
            status = "No confirmed active posting found"
            fill = "FFF2CC"
            roles = ""
            locations = ""
            urls = ""
            board = result.get("board")
            if board:
                urls = board.get("url", "")
                note = (
                    f"Live {board.get('provider')} board scanned "
                    f"({result.get('job_count', 0)} jobs); no job title matched internship/co-op/summer analyst."
                )
            elif result.get("official_source_domains"):
                urls = "; ".join(result.get("official_source_domains", []))
                note = (
                    "Official career source attempted; "
                    "no confirmed internship posting found."
                )
            else:
                note = (
                    "Could not verify via Greenhouse/Lever/Ashby automated board scan; "
                    "no confirmed official internship posting found."
                )
        values = [status, roles, locations, urls, note, scan_date]
        for col, output in enumerate(values, start_col):
            cell = sheet.cell(row, col, output)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == start_col:
                cell.fill = PatternFill("solid", fgColor=fill)

    summary_name = "Internship Scan Summary"
    if summary_name in workbook.sheetnames:
        del workbook[summary_name]
    summary = workbook.create_sheet(summary_name, 0)
    summary.append(
        ["Company", "Active Internship Found?", "Internship Role(s) Found", "Official Source URL(s)", "Scan Notes"]
    )
    for cell in summary[1]:
        style_header(cell)

    for row in range(2, sheet.max_row + 1):
        company = sheet.cell(row, 1).value
        if company and str(company).strip():
            summary.append(
                [
                    company,
                    sheet.cell(row, start_col).value,
                    sheet.cell(row, start_col + 1).value,
                    sheet.cell(row, start_col + 3).value,
                    sheet.cell(row, start_col + 4).value,
                ]
            )

    for row in summary.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row[1].value == "Yes":
            row[1].fill = PatternFill("solid", fgColor="C6EFCE")
        else:
            row[1].fill = PatternFill("solid", fgColor="FFF2CC")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:N{sheet.max_row}"
    for col, width in [(9, 24), (10, 70), (11, 45), (12, 85), (13, 85), (14, 16)]:
        sheet.column_dimensions[get_column_letter(col)].width = width

    summary.freeze_panes = "A2"
    summary.auto_filter.ref = f"A1:E{summary.max_row}"
    for col, width in [(1, 28), (2, 24), (3, 70), (4, 85), (5, 85)]:
        summary.column_dimensions[get_column_letter(col)].width = width

    workbook.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Scan a job tracker for active internships.")
    parser.add_argument("--input", required=True, help="Path to the source .xlsx tracker.")
    parser.add_argument("--output-dir", required=True, help="Directory for timestamped scan outputs.")
    parser.add_argument("--sheet", default="Tracking Template", help="Worksheet containing company names in column A.")
    parser.add_argument("--delay", type=float, default=0.1, help="Delay between company scans.")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y-%m-%d_%H%M")
    scan_date = dt.date.today().isoformat()
    output_path = os.path.join(args.output_dir, f"active_internship_scan_{stamp}.xlsx")
    raw_json_path = os.path.join(args.output_dir, f"active_internship_scan_{stamp}.json")

    companies = read_companies(args.input, args.sheet)
    results = []
    for index, (_, company) in enumerate(companies, 1):
        result = scan_company(company, args.delay)
        results.append(result)
        print(
            f"{index:02d}/{len(companies)} {company}: "
            f"{len(result['matches'])} internship match(es)"
        )

    write_workbook(args.input, output_path, args.sheet, results, scan_date)
    with open(raw_json_path, "w", encoding="utf-8") as handle:
        json.dump(results, handle, indent=2)

    yes_count = sum(1 for result in results if result.get("matches"))
    print("")
    print(f"Companies scanned: {len(results)}")
    print(f"Companies with confirmed internship postings: {yes_count}")
    print(f"Workbook: {output_path}")
    print(f"Raw scan JSON: {raw_json_path}")


if __name__ == "__main__":
    main()
